import os
import re
import time
import json
import shutil
import tempfile
import traceback
from datetime import datetime
from typing import List, Dict, Any, Optional
from urllib.parse import urlparse
import pandas as pd
import requests
import yaml
from dotenv import load_dotenv

# Selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException

# webdriver_manager (can be toggled/pinned via CHROMEDRIVER_VERSION)
from webdriver_manager.chrome import ChromeDriverManager

# Excel styling
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# ============================ ENV / GLOBALS ============================

load_dotenv()

def load_secrets_from_aws():
    """Fetch credentials from AWS Secrets Manager if not already set in env."""
    secret_name = "qa/an2p/healthcheck/login_credentials"
    region_name = "us-east-1"
    try:
        # Initialize AWS Secrets Manager client
        client = boto3.client("secretsmanager", region_name=region_name)
        get_secret_value_response = client.get_secret_value(SecretId=secret_name)
        secret_dict = json.loads(get_secret_value_response["SecretString"])

        # Populate environment variables only if not already present
        for key, value in secret_dict.items():
            os.environ.setdefault(key, value)

        print("✅ Loaded credentials from AWS Secrets Manager")
    except Exception as e:
        print(f"⚠️ Could not load secrets from AWS Secrets Manager: {e}")

# Call this before reading env vars
load_secrets_from_aws()
# --- 👆 End new block 👆 ---

def env_bool(name: str, default=False) -> bool:
    v = os.getenv(name, str(default)).strip().lower()
    return v in ("1", "true", "yes", "y", "on")

LOGIN_URL         = os.getenv("LOGIN_URL", "")
USERNAME          = os.getenv("OPSNOW_USERNAME", "")
PASSWORD          = os.getenv("OPSNOW_PASSWORD", "")
EXCEL_FILE_PREFIX = "global_health_check_report"
SLACK_WEBHOOK_URL = os.getenv("SLACK_WEBHOOK_URL", "")
HEADLESS          = env_bool("HEADLESS", True)
TIMEOUT           = int(os.getenv("TIMEOUT", "30"))
RENDER_RETRY      = int(os.getenv("RENDER_RETRY", "15"))
LOCATOR_DEBUG     = env_bool("LOCATOR_DEBUG", False)
CHROMEDRIVER_VERSION = os.getenv("CHROMEDRIVER_VERSION", "").strip()
CHROME_USER_DATA_DIR = os.getenv("CHROME_USER_DATA_DIR", "").strip()  # optional: use a fixed path

# Optional: fetch YAML from S3 instead of local file
CONFIG_URI        = ""  # Hardcoded to use local file
CONFIG_PATH_LOCAL = "global_config.yaml"

# Label keys for JS scan fallback
SERVER_LABEL_KEYS = ["total server", "total servers", "server", "servers", "서버", "총 서버"]


# ============================ FILE HELPERS ============================

driver: Optional[webdriver.Chrome] = None
wait:   Optional[WebDriverWait]    = None
TEMP_PROFILE_DIR: Optional[str]    = None 

def ts() -> str:
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

def ensure_dirs():
    os.makedirs("screenshots", exist_ok=True)
    os.makedirs("debug_html", exist_ok=True)
    os.makedirs("debug_json", exist_ok=True)

def safe_filename(name: str) -> str:
    if not isinstance(name, str):
        name = str(name)
    name = re.sub(r'[<>:"/\\|?*\n\r\t]+', "_", name)
    name = re.sub(r'__+', '_', name).strip("_ ")
    return name[:100] or "file"

def snap(name: str) -> str:
    ensure_dirs()
    path = os.path.join("screenshots", f"{safe_filename(name)}_{ts()}.png")
    try:
        if driver:
            driver.save_screenshot(path)
            print(f"Screenshot saved: {path}")
        else:
            # create an empty placeholder so caller has a path
            with open(path, "wb") as f:
                pass
            print(f"Driver not initialized; created placeholder screenshot: {path}")
    except Exception as e:
        print("Screenshot save failed:", e)
    return path

def dump_html(name: str) -> str:
    ensure_dirs()
    path = os.path.join("debug_html", f"{safe_filename(name)}_{ts()}.html")
    try:
        if driver:
            with open(path, "w", encoding="utf-8") as f:
                f.write(driver.page_source)
            print(f"HTML dumped: {path}")
        else:
            with open(path, "w", encoding="utf-8") as f:
                f.write("")
            print(f"Driver not initialized; created empty HTML dump: {path}")
    except Exception as e:
        print("HTML dump failed:", e)
    return path

def dump_json(name: str, obj: Any) -> str:
    ensure_dirs()
    path = os.path.join("debug_json", f"{safe_filename(name)}_{ts()}.json")
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(obj, f, ensure_ascii=False, indent=2)
        print(f"JSON dumped: {path}")
    except Exception as e:
        print("JSON dump failed:", e)
    return path


# ============================ SLACK HELPERS ============================

def _slack_text(check_name: str, incident_msg: str, screenshot_path: Optional[str]) -> str:
    file_name = os.path.basename(screenshot_path) if screenshot_path else ""
    return (
        "OpsNow360 Health Check\n"
        f"- Incident Message : [{check_name}] : Fail - \"{incident_msg}\"\n"
        f"- Add Screenshot : {file_name}"
    )

def slack_notify(check_name: str, incident_msg: str, screenshot_path: Optional[str]):
    try:
        if not SLACK_WEBHOOK_URL:
            return
        payload = {"text": _slack_text(check_name, incident_msg, screenshot_path)}
        r = requests.post(SLACK_WEBHOOK_URL, json=payload, timeout=10)
        if r.status_code >= 300:
            print(f"Slack webhook returned {r.status_code}: {r.text}")
    except Exception as e:
        print(f"Slack notify failed: {e}")


# ============================ SELENIUM/SSO ============================

def create_driver():
    """Create Chrome driver with a unique user-data-dir to avoid 'already in use' errors."""
    global TEMP_PROFILE_DIR

    chrome_options = Options()
    if HEADLESS:
        chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("--ignore-certificate-errors")
    chrome_options.add_argument("--allow-insecure-localhost")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--no-first-run")
    chrome_options.add_argument("--no-default-browser-check")
    chrome_options.add_argument("--disable-default-apps")
    chrome_options.add_argument("--disable-notifications")
    chrome_options.add_argument("--disable-features=TranslateUI")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option("useAutomationExtension", False)

    # --- KEY FIX: unique Chrome profile per run ---
    # If CHROME_USER_DATA_DIR env is set, use it; else create a temp dir and clean it up on exit.
    profile_dir = CHROME_USER_DATA_DIR
    if profile_dir:
        os.makedirs(profile_dir, exist_ok=True)
        TEMP_PROFILE_DIR = None  # user provided — don't delete
    else:
        TEMP_PROFILE_DIR = tempfile.mkdtemp(prefix="opsnow_chrome_")
        profile_dir = TEMP_PROFILE_DIR

    chrome_options.add_argument(f"--user-data-dir={profile_dir}")
    chrome_options.add_argument(f"--disk-cache-dir={os.path.join(profile_dir, 'cache')}")
    chrome_options.add_argument("--profile-directory=Default")

    # webdriver_manager
    if CHROMEDRIVER_VERSION:
        drv = webdriver.Chrome(
            service=Service(ChromeDriverManager(version=CHROMEDRIVER_VERSION).install()),
            options=chrome_options
        )
    else:
        drv = webdriver.Chrome(
            service=Service(ChromeDriverManager().install()),
            options=chrome_options
        )
    w = WebDriverWait(drv, TIMEOUT)
    return drv, w

def on_keycloak() -> bool:
    try:
        driver.find_element(By.NAME, "username")
        driver.find_element(By.NAME, "password")
        driver.find_element(By.ID, "kc-login")
        return True
    except NoSuchElementException:
        return False

def do_keycloak_login(current_url: Optional[str] = None,
                      username_override: Optional[str] = None,
                      password_override: Optional[str] = None):
    """
    Fill Keycloak credentials and submit.
    Priority:
      1) explicit username_override/password_override
      2) XERTICA_USERNAME/PASSWORD if current_url contains 'xertica'
      3) OPSNOW_USERNAME/PASSWORD
      4) global USERNAME/PASSWORD vars
    """
    if username_override is not None and password_override is not None:
        user = username_override
        pwd  = password_override
    else:
        if current_url and "xertica" in current_url.lower():
            user = os.getenv("XERTICA_USERNAME") or os.getenv("OPSNOW_USERNAME") or USERNAME
            pwd  = os.getenv("XERTICA_PASSWORD") or os.getenv("OPSNOW_PASSWORD") or PASSWORD
        else:
            user = os.getenv("OPSNOW_USERNAME") or USERNAME
            pwd  = os.getenv("OPSNOW_PASSWORD") or PASSWORD

    if not user or not pwd:
        raise ValueError("Keycloak credentials not found. "
                         "Set OPSNOW_USERNAME/OPSNOW_PASSWORD (and XERTICA_* for Xertica if needed).")

    try:
        u = wait.until(EC.presence_of_element_located((By.NAME, "username")))
        p = wait.until(EC.presence_of_element_located((By.NAME, "password")))
        b = wait.until(EC.element_to_be_clickable((By.ID, "kc-login")))
        u.clear(); u.send_keys(user)
        p.clear(); p.send_keys(pwd)
        b.click()
    except Exception:
        print("Exception while filling Keycloak login form:")
        traceback.print_exc()
        raise

def safe_do_keycloak_login(current_url: Optional[str] = None,
                           username_override: Optional[str] = None,
                           password_override: Optional[str] = None):
    try:
        do_keycloak_login(current_url=current_url,
                          username_override=username_override,
                          password_override=password_override)
    except Exception as e:
        print("safe_do_keycloak_login: failed:", e)
        traceback.print_exc()
        raise

def open_with_sso(url: str,
                  debug_name: str,
                  username_override: Optional[str] = None,
                  password_override: Optional[str] = None,
                  clear_cookies_before: bool = False):
    if clear_cookies_before:
        try:
            driver.delete_all_cookies()
        except Exception:
            pass
    driver.get(url)
    time.sleep(2)
    if on_keycloak():
        print(f"Keycloak detected — logging in for {url} …")
        safe_do_keycloak_login(current_url=url,
                               username_override=username_override,
                               password_override=password_override)
    try:
        WebDriverWait(driver, 15).until(
            lambda d: len(d.find_elements(By.CSS_SELECTOR,
                    "em.value, .value, .num, .number, .count, .am5-layer")) > 0
        )
    except TimeoutException:
        pass
    time.sleep(1)

def wait_non_empty_text(get_text, seconds=RENDER_RETRY) -> str:
    txt = ""
    for _ in range(seconds):
        try:
            txt = (get_text() or "").strip()
        except Exception:
            txt = ""
        if txt and txt != "0" and txt.lower() != "0ea":
            return txt
        time.sleep(1)
    return txt

def login_console():
    """Open LOGIN_URL and SSO in, waiting up to TIMEOUT for navigation."""
    if not LOGIN_URL:
        print("LOGIN_URL not set — skipping initial console login (will SSO per-page).")
        return
    print("Opening login page…", LOGIN_URL)
    driver.get(LOGIN_URL)
    time.sleep(2)
    if on_keycloak():
        print("Logging in to console via Keycloak…")
        do_keycloak_login(current_url=LOGIN_URL)
    try:
        host = (urlparse(LOGIN_URL).hostname or "")
        if host:
            wait.until(EC.url_contains(host))
    except Exception:
        pass
    time.sleep(1)
    print("Login successful (or skipped).")


# ============================ JS FALLBACKS ============================

def js_scan_labels() -> List[Dict[str, str]]:
    js = r"""
    const blocks = Array.from(document.querySelectorAll(
      ".count-item, .summary, .card, .cards, [class*=count], [class*=summary], [class*=kpi]"
    ));
    const rows = [];
    const pull = (root) => {
      const labelEl = root.querySelector("p, .label, .title, h3, h4, dt, .name");
      const valueEl = root.querySelector("em.value, .value, .num, .number, .count, dd em.value");
      const label = (labelEl && labelEl.textContent ? labelEl.textContent.trim() : "");
      const value = (valueEl && valueEl.textContent ? valueEl.textContent.trim() : "");
      if (value) rows.push({label, value});
    };
    if (blocks.length) {
      blocks.forEach(pull);
    } else {
      const vals = Array.from(document.querySelectorAll("em.value, .value, .num, .number, .count"));
      vals.forEach(v => {
        let node = v, label = "";
        for (let i = 0; i < 5 && node; i++) {
          const l = node.querySelector?.("p, .label, .title, h3, h4, dt, .name");
          if (l && l.textContent) { label = l.textContent.trim(); break; }
          node = node.parentElement;
        }
        const value = v.textContent.trim();
        if (value) rows.push({label, value});
      });
    }
    return rows;
    """
    return driver.execute_script(js) or []

def pick_value_by_labels(rows: List[Dict[str, str]], label_keys: List[str]) -> str:
    for r in rows:
        lab = (r.get("label") or "").strip().lower()
        val = (r.get("value") or "").strip()
        if not val:
            continue
        for key in label_keys:
            if key.lower() in lab:
                return val
    for r in rows:
        lab = (r.get("label") or "").strip().lower()
        val = (r.get("value") or "").strip()
        if lab == "" and val and val.replace(",", "").isdigit():
            return val
    return ""

def js_find_ec2_near_aws() -> str:
    js = r"""
    const isVisible = (el) => {
      if (!el) return false;
      const st = getComputedStyle(el);
      return st && st.display !== 'none' && st.visibility !== 'hidden' && (el.offsetParent !== null || el.getClientRects().length);
    };
    const all = Array.from(document.querySelectorAll("span,div,button,a,li,p"));
    const ec2s = all.filter(el => (el.textContent || "").trim().toLowerCase() === "ec2" && isVisible(el));
    for (const ec2 of ec2s) {
      let node = ec2;
      for (let i = 0; i < 6 && node; i++) {
        const txt = (node.textContent || "").toLowerCase();
        if (txt.includes("aws")) return "FOUND";
        node = node.parentElement;
      }
    }
    const any = all.find(el => (el.textContent || "").trim().toLowerCase() === "ec2" && isVisible(el));
    return any ? "FOUND" : "";
    """
    return driver.execute_script(js) or ""

def js_find_mtd_cost() -> str:
    js = r"""
    const money = t => /\$\s*[\d,]+(\.\d+)?/.test((t||"").trim());
    const cards = Array.from(document.querySelectorAll("*"))
      .filter(el => /month\s*to\s*date\s*cost/i.test(el.textContent || ""));
    for (const card of cards) {
      const spans = card.querySelectorAll("span.currency-text, span[class*=currency], span[class*=-number], span[class*=value]");
      for (const s of spans) {
        const t = (s.textContent || "").trim();
        if (money(t)) return t;
      }
    }
    const spans = Array.from(document.querySelectorAll("span.currency-text, span[class*=currency], span[class*=-number], span[class*=value]"));
    for (const s of spans) {
      const t = (s.textContent || "").trim();
      if (money(t)) return t;
    }
    return "";
    """
    return driver.execute_script(js) or ""

def js_find_more_available_total() -> str:
    js = r"""
    const money = t => /\$\s*[\d,]+(\.\d+)?/.test((t||"").trim());
    const sections = Array.from(document.querySelectorAll("section, div, article"))
      .filter(el => /more\s+available\s+cost\s+savings/i.test(el.textContent || ""));
    for (const sec of sections) {
      const live = Array.from(sec.querySelectorAll("article"))
        .find(a => !/display\s*:\s*none/i.test(a.getAttribute("style") || ""));
      const root = live || sec;
      const values = root.querySelectorAll("span, div, p, b, strong");
      for (const v of values) {
        const t = (v.textContent || "").trim();
        if (money(t)) return t;
      }
    }
    return "";
    """
    return driver.execute_script(js) or ""

def js_find_ceikpi_grade_full() -> str:
    js = r"""
    const sections = Array.from(document.querySelectorAll("section,div,article"))
      .filter(el => /total\s*scores/i.test(el.textContent || ""));
    const parenRe = /^\([\d.,]+\)$/;
    for (const sec of sections) {
      const blocks = Array.from(sec.querySelectorAll("p,div,span"))
        .filter(el => /grade/i.test((el.textContent || "")));
      for (const b of blocks) {
        const spans = Array.from(b.querySelectorAll("span"));
        let grade = "";
        let paren = "";
        for (const s of spans) {
          const t = (s.textContent || "").trim();
          if (/grade/i.test(t)) grade = t;
          if (parenRe.test(t)) paren = t;
        }
        if (grade) {
          return (grade + (paren ? " " + paren : "")).trim();
        }
      }
    }
    return "";
    """
    return driver.execute_script(js) or ""


# ============================ CONFIG LOADING ============================

def maybe_fetch_config_from_s3(uri: str, local: str = "config.yaml") -> str:
    if not uri:
        return local
    import boto3
    assert uri.startswith("s3://")
    bucket, key = uri[5:].split("/", 1)
    s3 = boto3.client("s3")
    s3.download_file(bucket, key, local)
    return local

def load_config(path: str) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        raw = yaml.safe_load(f) or {}
    # expected: version, defaults, metadata_by_url, checks
    raw.setdefault("defaults", {})
    raw.setdefault("metadata_by_url", {})
    raw.setdefault("checks", [])
    return raw


# ============================ REPORTING ============================

EXCEL_COLUMNS = [
    "Site","Company","Service","Menu","URL","Check","Locator","Value","Status","Screenshot"
]

def make_result(meta: Dict[str, str], url: str, check_name: str,
                locator_used: str, value: str, status: str,
                screenshot_path: Optional[str] = None) -> Dict[str, str]:
    if status == "FAIL" and not screenshot_path:
        screenshot_path = snap(check_name.replace(" ", "_") + "_Fail")
    screenshot_name = os.path.basename(screenshot_path) if screenshot_path else ""
    return {
        "Site": meta.get("Site", ""),
        "Company": meta.get("Company", ""),
        "Service": meta.get("Service", ""),
        "Menu": meta.get("Menu", ""),
        "URL": url,
        "Check": check_name,
        "Locator": locator_used or "",
        "Value": value or "",
        "Status": status,
        "Screenshot": screenshot_name
    }

def save_report(rows: List[Dict[str, str]]) -> str:
    df = pd.DataFrame(rows)
    for col in EXCEL_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    df = df[EXCEL_COLUMNS]

    out = f"{EXCEL_FILE_PREFIX}_{ts()}.xlsx"
    df.to_excel(out, index=False)

    wb = load_workbook(out)
    ws = wb.active
    header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    status_col = header_map.get("Status")

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    if status_col:
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=status_col)
            if cell.value == "PASS":
                cell.fill = green
            elif cell.value == "FAIL":
                cell.fill = red

    wb.save(out)
    print(f"\nReport saved: {out}")
    return out


# ============================ XERTICA HELPERS ============================

def switch_company_to_force(target_text="Xertica Clientes por reconocer",
                            wait_after=2, timeout=10) -> bool:
    """
    Try hard to switch the topbar/company selector to the target.
    Returns True if verified in topbar text.
    """
    target = target_text.strip().lower()

    def read_topbar_text() -> str:
        sel_candidates = [
            "div.header__company","div.company-name","a.topbar-company","button.company-toggle","div.header .company",
            "div.bs-select-inline",".company-selector",".header-company"
        ]
        for s in sel_candidates:
            try:
                el = driver.find_element(By.CSS_SELECTOR, s)
                if el and el.text:
                    return el.text.strip().lower()
            except Exception:
                pass
        try:
            els = driver.find_elements(By.CSS_SELECTOR, "header *, nav *, div.topbar *, div.header *")
            for e in els:
                t = (e.text or "").strip().lower()
                if t and len(t) < 120 and ("xertica" in t or "*" in t or "transportes" in t):
                    return t
        except Exception:
            pass
        return ""

    def poll_for_verify(seconds=6):
        for _ in range(seconds):
            cur = read_topbar_text()
            if cur and target in cur:
                return True, cur
            time.sleep(1)
        return False, read_topbar_text()

    try:
        ok, cur = poll_for_verify(seconds=1)
        if ok:
            print("Company already selected:", cur)
            return True

        js_switch = r"""
        const target = arguments[0].trim().toLowerCase();
        const isVisible = el => !!el && getComputedStyle(el).display !== 'none' && getComputedStyle(el).visibility !== 'hidden' && (el.offsetParent !== null || el.getClientRects().length);
        const toggles = Array.from(document.querySelectorAll('a,button,div,span,p')).filter(n=>{
          try{
            const t=(n.innerText||'').trim().toLowerCase();
            return isVisible(n) && (t.includes('*') || /transportes|company|empresa|cliente|client|8091/i.test(t) || n.getAttribute('aria-haspopup')==='true' || n.getAttribute('role')==='button');
          }catch(e){return false;}
        });
        if(toggles.length){
          try{ toggles[0].scrollIntoView({block:'center',inline:'center'}); toggles[0].click(); }catch(e){}
        }
        const findAndClick = ()=>{
          const opts = Array.from(document.querySelectorAll('li,div,button,a,span,p'))
            .filter(n=> isVisible(n) && (n.innerText||'').trim().toLowerCase().includes(target));
          if(opts.length){
            try{ opts[0].scrollIntoView({block:'center',inline:'center'}); opts[0].click(); return true;}catch(e){}
            try{
              opts[0].dispatchEvent(new MouseEvent('mousedown',{bubbles:true}));
              opts[0].dispatchEvent(new MouseEvent('mouseup',{bubbles:true}));
              opts[0].dispatchEvent(new MouseEvent('click',{bubbles:true}));
              return true;
            }catch(e){}
          }
          return false;
        };
        if(findAndClick()) return true;
        const end = Date.now() + 2000;
        while(Date.now() < end){
          if(findAndClick()) return true;
        }
        return false;
        """
        clicked = bool(driver.execute_script(js_switch, target))
        print("Company selector JS click attempted:", clicked)

        ok, cur = poll_for_verify(seconds=4)
        if ok:
            print("Switched company (verified):", cur)
            return True

        print("Company switch not verified; attempting reload as last resort…")
        driver.refresh()
        time.sleep(1.5)
        ok, cur = poll_for_verify(seconds=6)
        if ok:
            print("Switched company after reload (verified):", cur)
            return True
        else:
            print("After reload — still not switched. topbar:", cur)
    except Exception as e:
        print("switch_company_to_force error:", e)
    return False

def select_only_xertica_option(option_text="Xertica Clientes por reconocer") -> bool:
    """Light attempt to click an option with visible text (scheduler inner panel)."""
    try:
        txt = option_text.strip().lower()
        opt_xpath = f"//*[contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), {json.dumps(txt)})]"
        opts = driver.find_elements(By.XPATH, opt_xpath)
        for o in opts:
            if o.is_displayed():
                try:
                    o.click()
                    return True
                except Exception:
                    try:
                        driver.execute_script("arguments[0].click();", o)
                        return True
                    except Exception:
                        pass
        return False
    except Exception as e:
        print("select_only_xertica_option error:", e)
        return False


# ============================ CHECK DISPATCHER ============================

def run_one_check(check: Dict[str, Any], cfg: Dict[str, Any]) -> Dict[str, str]:
    """
    check: {
      name, url, type, locators:[{kind,value},...],
      js_fallback:{strategy:..., label_keys:[]}, metadata:{...},
      (optional) login_url, login_username, login_password
    }
    """
    url = check["url"]
    name = check["name"]
    ctype = (check.get("type") or "value_required").lower()
    locators = check.get("locators", [])
    js_fb = check.get("js_fallback") or {}
    per_meta = check.get("metadata") or {}

    # Build metadata precedence: defaults -> metadata_by_url[url] -> per-check metadata
    defaults = cfg.get("defaults", {})
    meta_by_url = cfg.get("metadata_by_url", {})
    meta: Dict[str, str] = {
        "Site": defaults.get("site", ""),
        "Company": defaults.get("company", ""),
        "Service": "",
        "Menu": "",
    }
    url_meta = meta_by_url.get(url, {})
    for k, v in url_meta.items():
        meta[k.capitalize()] = v
    for k, v in per_meta.items():
        meta[k.capitalize()] = v

    print(f"\n→ {name} @ {url}")

    # Optional per-check login first (if provided in YAML)
    login_url = check.get("login_url") or LOGIN_URL
    login_user = check.get("login_username")
    login_pass = check.get("login_password")

    # If this is a Xertica host and no explicit overrides, prefer XERTICA_* env if present
    if (login_url and "xertica" in login_url.lower()) and (login_user is None and login_pass is None):
        login_user = os.getenv("XERTICA_USERNAME") or login_user
        login_pass = os.getenv("XERTICA_PASSWORD") or login_pass

    # Try logging in to login_url (if set)
    try:
        if login_url:
            open_with_sso(login_url, name.replace(" ", "_") + "_login",
                          username_override=login_user, password_override=login_pass)
    except Exception as e:
        print(f"Login warning for {login_url}: {e} (continuing)")

    # --- NEW: also prefer Xertica creds for the target URL navigation if it's an xertica domain ---
    nav_user = None
    nav_pass = None
    if "xertica" in (url or "").lower():
        nav_user = os.getenv("XERTICA_USERNAME") or login_user
        nav_pass = os.getenv("XERTICA_PASSWORD") or login_pass
    else:
        # fall back to whatever overrides were set (or None)
        nav_user = login_user
        nav_pass = login_pass

    # Open target URL
    try:
        open_with_sso(url, name.replace(" ", "_"),
                      username_override=nav_user, password_override=nav_pass)
    except Exception as e:
        print(f"Navigation warning for {url}: {e} (continuing)")

    # Xertica scheduler-specific pre-steps (company switch + light option select)
    try:
        if "asset.xertica.cloud/asset/scheduler" in (url or "").lower() or \
           ("asset.xertica.cloud/asset/scheduler" in driver.current_url.lower()):
            try:
                switched = switch_company_to_force("Xertica Clientes por reconocer")
                print("Company switch attempted:", switched)
            except Exception as e:
                print("switch_company_to_force error (continuing):", e)
            try:
                tried = select_only_xertica_option("Xertica Clientes por reconocer")
                print("Xertica option selection attempted:", tried)
            except Exception as e:
                print("select_only_xertica_option error (continuing):", e)
    except Exception as e:
        print("Xertica pre-step error:", e)

    # Let SPA render something
    try:
        WebDriverWait(driver, 10).until(
            lambda d: len(d.find_elements(By.CSS_SELECTOR,
                    "em.value, .value, .num, .number, .count, .am5-layer")) > 0
        )
    except TimeoutException:
        pass

    # Optional deep locator debug
    if LOCATOR_DEBUG and locators:
        try:
            for loc in locators:
                kind = (loc.get("kind") or "xpath").lower()
                value = (loc.get("value") or "").strip()
                if not value:
                    continue
                try:
                    elems = (driver.find_elements(By.XPATH, value)
                             if kind == "xpath" else driver.find_elements(By.CSS_SELECTOR, value))
                    print(f"DEBUG locator ({kind}): {value}  ->  matches: {len(elems)}")
                    for i, e in enumerate(elems[:5]):
                        try:
                            outer = driver.execute_script("return arguments[0].outerHTML.slice(0,200);", e)
                        except Exception:
                            outer = "<outerHTML unavailable>"
                        try:
                            selenium_text = e.text
                        except Exception:
                            selenium_text = ""
                        try:
                            raw_text = driver.execute_script("return arguments[0].textContent;", e) or ""
                            inner_text = driver.execute_script("return arguments[0].innerText;", e) or ""
                        except Exception:
                            raw_text = inner_text = ""
                        print(f"   - match[{i}] displayed={e.is_displayed()} text={repr(selenium_text)}")
                        print(f"     textContent={repr(raw_text)}, innerText={repr(inner_text)}")
                        print(f"     outerHTML~200={outer}")
                except Exception as ex:
                    print("  DEBUG locator error for", value, ":", ex)
        except Exception:
            pass

    # Try primary locators
    elem = None
    matched = None
    for loc in locators:
        kind = (loc.get("kind") or "xpath").lower()
        value = (loc.get("value") or "").strip()
        if not value:
            continue
        try:
            cond = EC.visibility_of_element_located(
                (By.XPATH, value) if kind == "xpath" else (By.CSS_SELECTOR, value)
            )
            elem = WebDriverWait(driver, 10).until(cond)
            matched = f"{kind}:{value}"
            break
        except TimeoutException:
            continue

    # Specialized types
    if ctype == "mtd_cost":
        value = wait_non_empty_text(lambda: elem.text, 30) if elem else ""
        if not value or "$" not in value:
            print("Falling back to JS scan for Month to Date Cost…")
            value = js_find_mtd_cost().strip()
        status = "PASS" if value else "FAIL"
        res = make_result(meta, url, name, matched or "[JS MTD cost]", value, status)

    elif ctype == "more_available_total":
        value = (elem.text.strip() if elem else "") or js_find_more_available_total().strip()
        status = "PASS" if value else "FAIL"
        res = make_result(meta, url, name, matched or "[JS MoreAvailable Total]", value, status)

    elif ctype == "cei_grade":
        value = (wait_non_empty_text(lambda: elem.text, 15) if elem else "") or js_find_ceikpi_grade_full().strip()
        status = "PASS" if value else "FAIL"
        res = make_result(meta, url, name, matched or "[JS CEI grade]", value, status)

    elif ctype == "element_exists":
        status = "PASS" if elem else "FAIL"
        value = "FOUND" if elem else ""
        if status == "FAIL" and js_fb.get("strategy") == "ec2_near_aws":
            print("Falling back to JS: EC2 near AWS…")
            if js_find_ec2_near_aws():
                status, value, matched = "PASS", "FOUND", "[JS EC2-near-AWS]"
        res = make_result(meta, url, name, matched or "n/a", value, status)

    else:  # value_required (generic)
        value = wait_non_empty_text(lambda: elem.text, 15) if elem else ""
        if (not value) and js_fb.get("strategy") == "scan_labels":
            print("Falling back to JS scan for label-based value…")
            rows = js_scan_labels()
            dump_json("Asset_ScanLabels", rows)
            label_keys = js_fb.get("label_keys") or SERVER_LABEL_KEYS
            value = pick_value_by_labels(rows, label_keys)
        status = "PASS" if value else "FAIL"
        res = make_result(meta, url, name, matched or "[JS scan fallback]", value, status)

    if res["Status"] == "FAIL":
        reason = f"{ctype} failed (locator: {res['Locator']})"
        screenshot_full = os.path.join("screenshots", res["Screenshot"]) if res["Screenshot"] else None
        slack_notify(name, reason, screenshot_full)

    print(f"Result: {res['Status']} | Value: {res['Value']}")
    return res


# ============================ MAIN ============================

if __name__ == "__main__":
    print("Starting health check…")
    try:
        ensure_dirs()
        # Create driver up front so helpers can use it
        driver, wait = create_driver()

        # Optionally fetch config.yaml from S3
        cfg_path = maybe_fetch_config_from_s3(CONFIG_URI, CONFIG_PATH_LOCAL)
        cfg = load_config(cfg_path)

        # Allow YAML to override a couple of runtime defaults if provided
        TIMEOUT_yaml = cfg.get("defaults", {}).get("timeout")
        RENDER_yaml  = cfg.get("defaults", {}).get("render_retry")
        if TIMEOUT_yaml: TIMEOUT = int(TIMEOUT_yaml)
        if RENDER_yaml:  RENDER_RETRY = int(RENDER_yaml)

        # Initial console login (if LOGIN_URL is set)
        login_console()

        results: List[Dict[str, str]] = []
        for c in cfg.get("checks", []):
            try:
                results.append(run_one_check(c, cfg))
            except Exception as e:
                print(f"Check '{c.get('name','<unnamed>')}' crashed: {e}")
                try:
                    snap("Check_Crashed_" + safe_filename(c.get('name','unnamed')))
                    dump_html("Check_Crashed_" + safe_filename(c.get('name','unnamed')))
                except Exception:
                    pass
                # still record a FAIL row with minimal info
                meta = {"Site":"", "Company":"", "Service":"", "Menu":""}
                results.append(make_result(meta, c.get("url",""), c.get("name",""), "", "", "FAIL"))

        save_report(results)
        print("\nDone.")
    except Exception as e:
        print(f"Fatal error: {e}")
        try:
            snap("Fatal_Error")
            dump_html("Fatal_Error")
        except Exception:
            pass
        slack_notify("Fatal Error", str(e), None)
    finally:
        print("Closing browser…")
        try:
            if driver:
                driver.quit()
        except Exception:
            pass
        # Clean up the temp Chrome profile if we created one
        try:
            if TEMP_PROFILE_DIR and os.path.isdir(TEMP_PROFILE_DIR):
                shutil.rmtree(TEMP_PROFILE_DIR, ignore_errors=True)
        except Exception:
            pass

